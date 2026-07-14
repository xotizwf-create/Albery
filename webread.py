"""webread.py — shared helpers for reading documents and the public web from this box.

Why this module exists (2026-07-13, incident: employee sent a WB sales-funnel xlsx and two
wildberries.ru card links; the agent extracted 223 chars from the file and got blocked on WB):

- Marketplace/1C-style xlsx exports write rows without ``r=`` indexes and without a
  ``<dimension>``; openpyxl in read_only mode then sees only the first row of every sheet.
  ``extract_xlsx`` keeps the cheap read_only pass and falls back to a streaming XML parser
  (bounded memory, handles inline strings) whenever the first pass looks empty.
- All outbound traffic egresses through the Estonia VPN; many Russian sites (rbc.ru gives 401,
  marketplaces, gov sites) reject the foreign IP. ``direct_fetch`` retries one request from the
  physical Russian interface (eth0) without touching the VPN or global routing.
- wildberries.ru pages sit behind an ASN-level antibot (HTTP 498) that blocks this box from BOTH
  egress routes AND the r.jina.ai reader, while WB's static basket CDN serves the same card as
  JSON with no antibot. ``wb_card_result`` answers card links from the CDN.

Kill switches: FETCH_URL_DIRECT=0 disables direct_fetch, FETCH_URL_WB_CARD=0 disables the WB
CDN handler; FETCH_URL_DIRECT_IFACE overrides the direct interface (default eth0).
"""
from __future__ import annotations

import gzip
import io
import json
import logging
import os
import re
import socket
import time
import urllib.parse
import urllib.request
import zipfile

log = logging.getLogger(__name__)

_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")


# --------------------------------------------------------------------------------------------
# xlsx extraction
# --------------------------------------------------------------------------------------------

_XLSX_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XLSX_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _xlsx_openpyxl(data: bytes, max_chars: int) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    total = 0
    for ws in wb.worksheets:
        out.append("# Лист: " + str(ws.title))
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                line = " | ".join(cells)
                out.append(line)
                total += len(line)
                if total > max_chars:
                    out.append("…[дальше обрезано: лимит извлечения]")
                    return "\n".join(out).strip()
    return "\n".join(out).strip()


def _xlsx_stream(data: bytes, max_chars: int) -> str:
    """Stream-parse sheet XML directly. Unlike openpyxl's read_only mode this does not trust
    row indexes / dimensions, so exports that omit them (WB, some 1C) are read in full."""
    from xml.etree.ElementTree import fromstring, iterparse

    z = zipfile.ZipFile(io.BytesIO(data))
    names = set(z.namelist())

    shared: list[str] = []
    if "xl/sharedStrings.xml" in names:
        for _ev, el in iterparse(z.open("xl/sharedStrings.xml")):
            if el.tag == _XLSX_MAIN_NS + "si":
                shared.append("".join(t.text or "" for t in el.iter(_XLSX_MAIN_NS + "t")))
                el.clear()

    # Sheet display names in workbook order -> worksheet part paths (via the rels map).
    rels: dict[str, str] = {}
    if "xl/_rels/workbook.xml.rels" in names:
        for rel in fromstring(z.read("xl/_rels/workbook.xml.rels")):
            rels[rel.get("Id") or ""] = rel.get("Target") or ""
    sheets: list[tuple[str, str]] = []
    if "xl/workbook.xml" in names:
        root = fromstring(z.read("xl/workbook.xml"))
        for sh in root.iter(_XLSX_MAIN_NS + "sheet"):
            target = rels.get(sh.get(_XLSX_REL_NS + "id") or "", "")
            target = target.lstrip("/")
            if target and not target.startswith("xl/"):
                target = "xl/" + target
            if target in names:
                sheets.append((sh.get("name") or "Лист", target))
    if not sheets:
        sheets = [(p.rsplit("/", 1)[-1], p) for p in sorted(names) if p.startswith("xl/worksheets/sheet")]

    out: list[str] = []
    total = 0
    for title, path in sheets:
        out.append("# Лист: " + title)
        for _ev, el in iterparse(z.open(path)):
            if el.tag != _XLSX_MAIN_NS + "row":
                continue
            cells: list[str] = []
            for c in el.iter(_XLSX_MAIN_NS + "c"):
                ctype = c.get("t") or ""
                val = ""
                if ctype == "inlineStr":
                    val = "".join(t.text or "" for t in c.iter(_XLSX_MAIN_NS + "t"))
                else:
                    v = c.find(_XLSX_MAIN_NS + "v")
                    val = (v.text or "") if v is not None else ""
                    if ctype == "s":
                        try:
                            val = shared[int(val)]
                        except (ValueError, IndexError):
                            pass
                cells.append(val)
            el.clear()
            if any(x.strip() for x in cells):
                line = " | ".join(cells)
                out.append(line)
                total += len(line)
                if total > max_chars:
                    out.append("…[дальше обрезано: лимит извлечения]")
                    return "\n".join(out).strip()
    return "\n".join(out).strip()


def extract_xlsx(data: bytes, max_chars: int = 600_000) -> str:
    """Full text of an xlsx/xlsm: cheap openpyxl read_only pass first, streaming-XML fallback
    when the result looks empty (missing row indexes) or openpyxl fails outright."""
    text = ""
    try:
        text = _xlsx_openpyxl(data, max_chars)
    except Exception as exc:  # noqa: BLE001
        log.warning("extract_xlsx: openpyxl pass failed: %s", repr(exc)[:160])
    if len(text) < 1500:
        try:
            alt = _xlsx_stream(data, max_chars)
            if len(alt) > len(text):
                text = alt
        except Exception as exc:  # noqa: BLE001
            log.warning("extract_xlsx: stream pass failed: %s", repr(exc)[:160])
    return text


# --------------------------------------------------------------------------------------------
# Direct (non-VPN) fetch from the Russian interface
# --------------------------------------------------------------------------------------------

_DOC_CTYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_DOC_URL_EXT_RE = re.compile(r"\.(pdf|docx|xlsx|xlsm)($|[?#])", re.I)


def direct_fetch(url: str, headers: dict | None = None, timeout: int = 25,
                 doc_limit: int = 12 * 1024 * 1024, text_limit: int = 350_000):
    """GET ``url`` with sockets bound to the physical RU interface (SO_BINDTODEVICE), bypassing
    the VPN default route. Returns (status, content_type, body, final_url) on 2xx, else None.
    Follows redirects; every hop stays on the direct interface."""
    if os.getenv("FETCH_URL_DIRECT", "1").strip().lower() in {"0", "false", "no", "off"}:
        return None
    iface = (os.getenv("FETCH_URL_DIRECT_IFACE", "eth0").strip() or "eth0").encode() + b"\0"
    import http.client
    import ssl

    class _HTTPS(http.client.HTTPSConnection):
        def connect(self):  # noqa: D102
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                self.sock.setsockopt(socket.SOL_SOCKET, 25, iface)  # 25 = SO_BINDTODEVICE
            except OSError:
                pass
            self.sock.settimeout(self.timeout)
            self.sock.connect((self.host, self.port))
            ctx = ssl.create_default_context()
            self.sock = ctx.wrap_socket(self.sock, server_hostname=self.host)

    class _HTTP(http.client.HTTPConnection):
        def connect(self):  # noqa: D102
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                self.sock.setsockopt(socket.SOL_SOCKET, 25, iface)
            except OSError:
                pass
            self.sock.settimeout(self.timeout)
            self.sock.connect((self.host, self.port))

    class _HsHandler(urllib.request.HTTPSHandler):
        def https_open(self, req):  # noqa: D102
            return self.do_open(_HTTPS, req)

    class _HHandler(urllib.request.HTTPHandler):
        def http_open(self, req):  # noqa: D102
            return self.do_open(_HTTP, req)

    try:
        opener = urllib.request.build_opener(_HHandler(), _HsHandler())
        req = urllib.request.Request(url, headers=dict(headers or {"User-Agent": _UA}))
        with opener.open(req, timeout=timeout) as resp:
            status = getattr(resp, "status", 200)
            ct = resp.headers.get("Content-Type", "") or ""
            is_doc = ct.split(";")[0].strip().lower() in _DOC_CTYPES or bool(_DOC_URL_EXT_RE.search(url))
            body = resp.read(doc_limit if is_doc else text_limit)
            if not (200 <= status < 300):
                return None
            return status, ct, body, (resp.geturl() or url)
    except Exception as exc:  # noqa: BLE001
        log.info("direct_fetch failed for %s: %s", url.split("?")[0][:120], repr(exc)[:160])
        return None


# --------------------------------------------------------------------------------------------
# Wildberries card via the public basket CDN
# --------------------------------------------------------------------------------------------

_WB_URL_RE = re.compile(r"wildberries\.(?:ru|by|kz)/catalog/(\d{5,12})", re.I)
# vol -> basket host number, learned at runtime; the anchors below are rough interpolation
# seeds (observed 2026-07: vol 7910 -> basket-36, vol 9695 -> basket-41).
_wb_basket_cache: dict[int, int] = {}
_WB_ANCHORS = [(0, 1), (1500, 8), (3500, 18), (5500, 26), (7910, 36), (9695, 41), (12000, 52)]
_WB_MAX_BASKET = 90
_WB_MAX_PROBES = 30


def _wb_get_json(url: str, timeout: int = 6):
    req = urllib.request.Request(url, headers={"User-Agent": _UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = resp.read(2 * 1024 * 1024)
    if body[:2] == b"\x1f\x8b":
        body = gzip.decompress(body)
    return json.loads(body)


def _wb_probe_order(vol: int) -> list[int]:
    anchors = sorted(set(_WB_ANCHORS) | set(_wb_basket_cache.items()))
    lo = max((a for a in anchors if a[0] <= vol), default=None)
    hi = min((a for a in anchors if a[0] > vol), default=None)
    if lo and hi and hi[0] > lo[0]:
        est = round(lo[1] + (hi[1] - lo[1]) * (vol - lo[0]) / (hi[0] - lo[0]))
    elif lo:
        est = lo[1] + max(0, (vol - lo[0]) // 450)  # ~450 vols per basket at the top end
    else:
        est = 1
    est = min(max(est, 1), _WB_MAX_BASKET)
    order: list[int] = []
    for d in range(_WB_MAX_BASKET):
        for cand in (est + d, est - d):
            if 1 <= cand <= _WB_MAX_BASKET and cand not in order:
                order.append(cand)
    return order


_WB_MISS_TTL_S = 600.0
_wb_card_miss: dict[int, float] = {}    # nm -> monotonic expiry (card not on CDN)
_wb_price_miss: dict[int, float] = {}   # nm -> monotonic expiry (no live price in search)
_WB_HIT_TTL_S = float(os.getenv("WB_PRICE_HIT_TTL_S", "2700") or "2700")  # 45 min
_wb_price_hit: dict[int, dict] = {}     # nm -> {min,max,sizes_priced,rating,feedbacks,exp}
_wb_route_toggle = [0]
_wb_absent: set = set()   # nm confirmed NOT returned by search on any query (out of stock / hidden)


def _wb_price_from_product(p: dict) -> dict:
    prices = []
    for s in (p.get("sizes") or []):
        pr = ((s.get("price") or {}).get("product") or (s.get("price") or {}).get("total") or 0)
        if pr:
            prices.append(round(pr / 100))
    return {
        "min": min(prices) if prices else None,
        "max": max(prices) if prices else None,
        "sizes_priced": len(prices),
        "rating": p.get("reviewRating") or p.get("nmReviewRating") or p.get("rating"),
        "feedbacks": p.get("feedbacks") or p.get("nmFeedbacks"),
    }


def _wb_cache_products(products: list) -> None:
    """Cache the price of EVERY product a search returned (a search = a catalog page). Sofya's
    list is dominated by siblings that co-appear, so one search prices many requested articles."""
    now = time.monotonic()
    for p in products or []:
        try:
            nm = int(p.get("id") or 0)
        except (TypeError, ValueError):
            continue
        if not nm:
            continue
        info = _wb_price_from_product(p)
        if info["min"] or info["rating"]:
            info["exp"] = now + _WB_HIT_TTL_S
            _wb_price_hit[nm] = info


def _wb_cached_price(nm: int):
    info = _wb_price_hit.get(int(nm))
    if info and info.get("exp", 0) > time.monotonic() and info.get("min"):
        return {k: info[k] for k in ("min", "max", "sizes_priced", "rating", "feedbacks")}
    return None


def _wb_fetch_card(nm: int, deadline: float | None = None):
    exp = _wb_card_miss.get(nm)
    if exp and exp > time.monotonic():
        return None, None  # known miss — answer instantly instead of re-sweeping hosts
    vol, part = nm // 100_000, nm // 1_000
    for b in _wb_probe_order(vol)[:_WB_MAX_PROBES]:
        if deadline is not None and time.monotonic() > deadline:
            return None, None  # deadline, NOT a confirmed miss — do not poison the cache
        host = "basket-%02d.wbbasket.ru" % b
        try:
            card = _wb_get_json(f"https://{host}/vol{vol}/part{part}/{nm}/info/ru/card.json", 4)
        except Exception:  # noqa: BLE001
            continue
        _wb_basket_cache[vol] = b
        return host, card
    _wb_card_miss[nm] = time.monotonic() + _WB_MISS_TTL_S
    return None, None


_WB_SEARCH_MIN_INTERVAL_S = float(os.getenv("WB_SEARCH_MIN_INTERVAL_S", "2.0") or "2.0")
_wb_last_search_ts = [0.0]


_wb_search_throttled_until = [0.0]


def wb_search_throttled() -> bool:
    """True while search.wb.ru is rate-limiting us — callers report an honest «WB тротлит,
    повторите через пару минут» instead of silently returning «цены нет»."""
    return time.monotonic() < _wb_search_throttled_until[0]


def _wb_search_direct_json(url: str, timeout: int):
    """Search via the DIRECT Russian interface: its 429-quota is separate from the VPN egress
    (which every other outbound request shares and burns), and a RU IP is the natural client
    for a RU marketplace. Returns parsed json or None (any failure -> caller falls back to VPN)."""
    d = direct_fetch(url, {"User-Agent": _UA, "Accept": "*/*"}, timeout=timeout,
                     text_limit=3 * 1024 * 1024)
    if not d:
        return None
    body = d[2]
    if body[:2] == b"\x1f\x8b":
        body = gzip.decompress(body)
    try:
        return json.loads(body)
    except Exception:  # noqa: BLE001
        return None


def _wb_search_paced(url: str, timeout: int = 10, deadline: float | None = None):
    """search.wb.ru rate-limits per IP (429). We ALTERNATE both egress IPs (direct-RU + VPN =
    two independent 429 buckets) and back off briefly on 429 — WITHOUT any global cooldown, so a
    hot moment never nukes the rest of the batch. ``deadline`` bounds the dance for one element."""
    for attempt in range(4):
        wait = _wb_last_search_ts[0] + _WB_SEARCH_MIN_INTERVAL_S - time.time()
        if deadline is not None and time.monotonic() + max(wait, 0) > deadline:
            raise TimeoutError("wb search: element deadline")
        if wait > 0:
            time.sleep(wait)
        _wb_last_search_ts[0] = time.time()
        t = timeout
        if deadline is not None:
            t = max(2, min(timeout, int(deadline - time.monotonic())))
        order = ("direct", "vpn") if _wb_route_toggle[0] % 2 == 0 else ("vpn", "direct")
        _wb_route_toggle[0] += 1
        for route in order:
            try:
                if route == "direct":
                    j = _wb_search_direct_json(url, t)
                    if j is not None:
                        return j
                else:
                    return _wb_get_json(url, t)
            except urllib.error.HTTPError as exc:
                if exc.code != 429:
                    raise
            except Exception:  # noqa: BLE001
                pass
        backoff = min(2.5 * (attempt + 1), 8.0)
        if deadline is not None and time.monotonic() + backoff > deadline:
            raise TimeoutError("wb search: element deadline (429)")
        time.sleep(backoff)
    raise TimeoutError("wb search: exhausted")


def _wb_current_prices(nm: int, card: dict, deadline: float | None = None):
    """Live storefront prices for a card via search.wb.ru (NOT antibot-blocked, unlike
    card.wb.ru): search by brand+name / vendor code, match the nm in results, read
    sizes[].price.product. Returns dict {min, max, sizes_priced, rating, feedbacks} in RUB
    or None. This is the price the customer sees BEFORE the personal WB-wallet discount."""
    name = str(card.get("imt_name") or "").strip()
    brand = str((card.get("selling") or {}).get("brand_name") or "").strip()
    vendor = str(card.get("vendor_code") or "").strip()
    # brand+name + vendor_code are enough; the name-only query was noise that made confirming
    # a genuine miss too slow to ever fit the per-item budget (so misses never got cached).
    queries = [q for q in (f"{brand} {name}".strip(), vendor) if q]
    # A card with no stock in the default region (Москва) is HIDDEN from its search results
    # while other regions still return it — try a second dest before giving up.
    dests = [d.strip() for d in os.getenv("WB_SEARCH_DESTS", "-1257786,-2162196").split(",") if d.strip()]
    cached = _wb_cached_price(nm)
    if cached:
        return cached
    exp = _wb_price_miss.get(nm)
    if exp and exp > time.monotonic():
        return None  # known miss — don't repeat the whole failing dance
    if deadline is None:
        deadline = time.monotonic() + float(os.getenv("WB_CARD_ELEMENT_BUDGET_S", "25") or "25")
    combos = [(q, d) for q in queries for d in dests]
    deadline_hit = False
    for query, dest in combos:
        if time.monotonic() > deadline:
            deadline_hit = True
            break
        try:
            url = ("https://search.wb.ru/exactmatch/ru/common/v5/search?appType=1&curr=rub"
                   f"&dest={dest}&resultset=catalog&suppressSpellcheck=false&query="
                   + urllib.parse.quote(query))
            j = _wb_search_paced(url, deadline=deadline)
        except TimeoutError:
            deadline_hit = True
            break
        except Exception:  # noqa: BLE001
            continue
        products = j.get("products") or (j.get("data") or {}).get("products") or []
        _wb_cache_products(products)  # price siblings for free — persists across calls/turns
        hit = next((p for p in products if int(p.get("id") or 0) == int(nm)), None)
        if not hit:
            continue
        prices = []
        for s in (hit.get("sizes") or []):
            p = ((s.get("price") or {}).get("product")
                 or (s.get("price") or {}).get("total") or 0)
            if p:
                prices.append(round(p / 100))
        out = {
            "min": min(prices) if prices else None,
            "max": max(prices) if prices else None,
            "sizes_priced": len(prices),
            "rating": hit.get("reviewRating") or hit.get("nmReviewRating") or hit.get("rating"),
            "feedbacks": hit.get("feedbacks") or hit.get("nmFeedbacks"),
        }
        if out["min"] or out["rating"]:
            _wb_absent.discard(int(nm))
            return out
    if deadline_hit:
        # ran out of the element budget — remember it briefly so an immediate retry (same task,
        # same list) answers instantly instead of chewing the same slow article again
        _wb_price_miss[nm] = time.monotonic() + float(os.getenv("WB_SLOW_MISS_TTL_S", "180") or "180")
    else:
        # every combo genuinely answered «нет такой карточки в выдаче» — confirmed absent
        _wb_price_miss[nm] = time.monotonic() + _WB_MISS_TTL_S
        _wb_absent.add(int(nm))
    return None


def wb_bulk_prices(articles: list[int], call_budget_s: float | None = None,
                   per_item_budget_s: float | None = None) -> dict:
    """Current storefront prices for a LIST of nm ids — the fast path for «заполни таблицу цен».
    One unreachable article can no longer eat the turn: each item gets its own deadline, known
    misses answer from cache, and when the call budget runs out the REST is returned in
    ``skipped`` so the caller just calls again with them. Never raises."""
    call_budget_s = call_budget_s or float(os.getenv("WB_BULK_CALL_BUDGET_S", "150") or "150")
    per_item_budget_s = per_item_budget_s or float(os.getenv("WB_BULK_ITEM_BUDGET_S", "22") or "22")
    t0 = time.monotonic()
    rows, skipped = [], []
    for nm in articles:
        try:
            nm = int(nm)
        except (TypeError, ValueError):
            rows.append({"article": nm, "status": "bad_article"})
            continue
        cached = _wb_cached_price(nm)  # free hit: this run already fetched it or a sibling search did
        if cached and cached.get("min"):
            rows.append({"article": nm, "status": "ok", "price_min": cached["min"],
                         "price_max": cached["max"], "sizes_priced": cached.get("sizes_priced")})
            continue
        if time.monotonic() - t0 > call_budget_s:
            skipped.append(nm)
            continue
        deadline = min(time.monotonic() + per_item_budget_s, t0 + call_budget_s + 5)
        row = {"article": nm, "status": "no_price", "price_min": None, "price_max": None}
        try:
            host, card = _wb_fetch_card(nm, deadline=deadline)
            if not card:
                row["status"] = "card_hidden"
            else:
                row["name"] = str(card.get("imt_name") or "")[:80]
                live = _wb_current_prices(nm, card, deadline=deadline)
                if live and live.get("min"):
                    row.update(status="ok", price_min=live["min"], price_max=live["max"],
                               sizes_priced=live.get("sizes_priced"))
                elif int(nm) in _wb_absent:
                    row["status"] = "card_hidden"  # WB doesn't return it in search — genuinely no price
        except Exception as exc:  # noqa: BLE001
            log.warning("wb_bulk_prices: nm=%s failed: %s", nm, repr(exc)[:120])
            row["status"] = "error"
        rows.append(row)
    return {
        "rows": rows,
        "skipped_by_budget": skipped,
        "ok": sum(1 for r in rows if r.get("status") == "ok"),
        "note": ("status=ok — цена сейчас на витрине (без скидки WB-кошелька); внеси сразу. "
                 "card_hidden — WB скрыл карточку из выдачи (нет в наличии), это единственный случай, "
                 "когда цены реально нет — так и скажи по этим артикулам. no_price/error/"
                 "skipped_by_budget — НЕ финал: WB просто ограничил частоту, вызови get_wb_prices ещё "
                 "раз с ЭТИМИ же артикулами (успешные уже в кэше, ответит быстро) — так добери все. "
                 "Повторяй, пока не останутся только card_hidden. Частичный результат вноси сразу."),
    }


def wb_card_result(url: str, max_chars: int = 50_000):
    """fetch_url-shaped result for a wildberries.ru card link, built from the basket CDN
    (card.json + price-history.json + feedbacks). None when the URL is not a WB card or the
    CDN gave nothing — the caller then falls through to the normal fetch path."""
    if os.getenv("FETCH_URL_WB_CARD", "1").strip().lower() in {"0", "false", "no", "off"}:
        return None
    m = _WB_URL_RE.search(url)
    if not m:
        return None
    nm = int(m.group(1))
    host, card = _wb_fetch_card(nm)
    if not card:
        return None
    vol, part = nm // 100_000, nm // 1_000

    lines: list[str] = []
    name = str(card.get("imt_name") or card.get("slug") or "").strip() or f"артикул {nm}"
    lines.append(f"# Wildberries — {name} (артикул {nm})")
    selling = card.get("selling") or {}
    lines.append("Бренд: %s | Категория: %s / %s | Артикул продавца: %s" % (
        selling.get("brand_name") or "?", card.get("subj_root_name") or "?",
        card.get("subj_name") or "?", card.get("vendor_code") or "?"))

    live = None
    try:
        live = _wb_current_prices(nm, card)
    except Exception:  # noqa: BLE001
        log.warning("wb card: live price lookup failed nm=%s", nm, exc_info=True)
    if live and live.get("min"):
        if live["min"] == live["max"]:
            lines.append(f"ЦЕНА СЕЙЧАС: {live['min']} ₽ — одинакова для всех размеров в наличии "
                         f"({live['sizes_priced']} шт). Это витринная цена без персональной скидки WB-кошелька.")
        else:
            lines.append(f"ЦЕНА СЕЙЧАС: от {live['min']} до {live['max']} ₽ в зависимости от размера "
                         f"({live['sizes_priced']} размеров в наличии). Витринная цена без скидки WB-кошелька.")
    if live and live.get("rating"):
        lines.append(f"Рейтинг: {live['rating']}/5, отзывов: {live.get('feedbacks') or '?'} (живые данные выдачи)")
    try:
        ph = _wb_get_json(f"https://{host}/vol{vol}/part{part}/{nm}/info/price-history.json", 5)
    except Exception:  # noqa: BLE001
        ph = None
    if ph:
        import datetime as _dt
        prices = [(p.get("price", {}).get("RUB") or 0) / 100 for p in ph]
        last = ph[-1]
        when = _dt.datetime.utcfromtimestamp(int(last.get("dt") or 0)).strftime("%d.%m.%Y")
        lines.append(
            f"Динамика цены во времени (история изменений, {len(ph)} точек, последняя {when}): "
            f"от {min(prices):g} до {max(prices):g} ₽. ⚠️ Это изменения цены ПО ДАТАМ — "
            f"НЕ разброс «цена от/до» по размерам, для таблиц цен используй ЦЕНУ СЕЙЧАС выше.")
    if not (live and live.get("min")) and not ph:
        lines.append("Цена: получить не удалось — попроси скриншот карточки, если цена критична.")
    elif not (live and live.get("min")):
        lines.append("⚠️ Актуальную витринную цену получить не удалось. НЕ выдавай исторические "
                     "мин/макс за «цену от/до» — честно скажи, что текущей цены нет.")

    imt = card.get("imt_id")
    if imt and not (live and live.get("rating")):
        fb = None
        for fh in ("feedbacks1.wb.ru", "feedbacks2.wb.ru"):
            try:
                fb = _wb_get_json(f"https://{fh}/feedbacks/v1/{imt}", 6)
            except Exception:  # noqa: BLE001
                continue
            if fb and (fb.get("valuation") or fb.get("feedbackCount")):
                break
        if fb and (fb.get("valuation") or fb.get("feedbackCount")):
            lines.append("Рейтинг: %s/5, отзывов: %s" % (
                fb.get("valuation") or "?", fb.get("feedbackCount") or fb.get("feedbackCountWithText") or "?"))

    colors = card.get("nm_colors_names")
    if colors:
        lines.append("Цвета: " + str(colors))
    media = card.get("media") or {}
    photo_count = media.get("photo_count") or 0
    lines.append(f"Фото в карточке: {photo_count} шт. Первое фото: "
                 f"https://{host}/vol{vol}/part{part}/{nm}/images/big/1.webp")

    opts = card.get("options") or []
    if opts:
        lines.append("\n## Характеристики")
        for o in opts:
            lines.append(f"- {o.get('name')}: {o.get('value')}")
    desc = str(card.get("description") or "").strip()
    if desc:
        lines.append("\n## Описание\n" + desc)
    lines.append(f"\nСтраница: https://www.wildberries.ru/catalog/{nm}/detail.aspx")

    text = "\n".join(lines)
    truncated = len(text) > max_chars
    return {
        "ok": True, "original_url": url,
        "fetched_url": f"https://{host}/vol{vol}/part{part}/{nm}/info/ru/card.json",
        "final_url": url, "kind": "wb-card", "status": 200,
        "content_type": "text/markdown", "char_count": min(len(text), max_chars),
        "truncated": truncated, "text": text[:max_chars],
        "note": ("Карточка получена из открытого CDN Wildberries: страницы wildberries.ru закрыты "
                 "антиботом для серверов. Точную витринную цену со скидкой WB-кошелька, остатки и "
                 "позиции в выдаче отсюда не видно — при необходимости попроси скриншот."),
    }
